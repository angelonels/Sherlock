from __future__ import annotations

from io import BytesIO
from typing import Any
from zipfile import BadZipFile, ZipFile

from openpyxl import load_workbook

from app.core.config import Settings
from app.services.column_cleaner import build_detected_columns, json_safe_value
from app.services.upload_safety import upload_error, warning_for_rows


def inspect_xlsx_zip(content: bytes, settings: Settings) -> None:
    try:
        with ZipFile(BytesIO(content)) as workbook_zip:
            total_uncompressed = sum(info.file_size for info in workbook_zip.infolist())
            total_compressed = max(sum(info.compress_size for info in workbook_zip.infolist()), 1)
    except BadZipFile as exc:
        raise upload_error("INVALID_XLSX", "XLSX file is not a valid workbook.") from exc

    if total_uncompressed > settings.xlsx_max_uncompressed_bytes:
        raise upload_error("XLSX_TOO_LARGE", "XLSX workbook expands beyond the configured safety limit.")
    if total_uncompressed / total_compressed > settings.xlsx_max_compression_ratio:
        raise upload_error("SUSPICIOUS_XLSX", "XLSX workbook has a suspicious compression ratio.")


def inspect_xlsx(content: bytes, settings: Settings, selected_sheet_name: str | None = None) -> dict[str, Any]:
    inspect_xlsx_zip(content, settings)
    workbook = load_workbook(BytesIO(content), read_only=True, data_only=True)
    sheet_names = list(workbook.sheetnames)
    if not sheet_names:
        raise upload_error("EMPTY_UPLOAD", "XLSX workbook does not contain any sheets.")

    recommended_sheet_name = sheet_names[0]
    sheet_name = selected_sheet_name or recommended_sheet_name
    if sheet_name not in sheet_names:
        raise upload_error("SHEET_NOT_FOUND", "Selected sheet was not found in the workbook.")

    sheet = workbook[sheet_name]
    row_iter = sheet.iter_rows(values_only=True)
    try:
        raw_headers = next(row_iter)
    except StopIteration as exc:
        raise upload_error("EMPTY_UPLOAD", "Selected XLSX sheet is empty.") from exc

    headers = [str(value).strip() if value not in (None, "") else f"Column {index}" for index, value in enumerate(raw_headers, start=1)]
    if len(headers) > settings.upload_max_columns:
        raise upload_error("TOO_MANY_COLUMNS", "Uploaded file has too many columns.", {"max_columns": settings.upload_max_columns})

    preview_rows: list[dict[str, Any]] = []
    for raw_row in row_iter:
        row_values = list(raw_row[: len(headers)])
        if not any(value not in (None, "") for value in row_values):
            continue
        preview_rows.append({header: json_safe_value(value) for header, value in zip(headers, row_values, strict=False)})
        if len(preview_rows) >= settings.upload_preview_rows:
            break

    if not preview_rows:
        raise upload_error("HEADERS_ONLY_UPLOAD", "Selected XLSX sheet contains headers but no data rows.")

    return {
        "sheet_names": sheet_names,
        "selected_sheet_name": selected_sheet_name,
        "recommended_sheet_name": recommended_sheet_name,
        "preview_rows": preview_rows,
        "detected_columns": build_detected_columns(headers, preview_rows),
        "warnings": warning_for_rows(preview_rows, settings),
    }
